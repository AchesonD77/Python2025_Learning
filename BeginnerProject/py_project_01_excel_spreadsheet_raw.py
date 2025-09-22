# python program can process thousands of spreadsheets in under a second
# excel Spreadsheets

# 1. import packages and make a short name,
# before doing this we need to go to PYPi to download the package - openpyxl, or just use code: pip install openpyxl on terminal
import openpyxl as xl

# 2. call a fuction, return a wordbook object
wb = xl.load_workbook('transactions.xlsx')
# 3. go to access your Sheet1 by (capital it), return our Sheet1
sheet = wb['Sheet1']
# 4. call the cell in our sheet, just use the coordinate 'a1' 'a2'
cell = sheet['a1']
print(cell.value)
# another approach
cell = sheet.cell(1,1)
print(cell.value)

# 5. method to know how many rows we have
print(sheet.max_row)

# 6. range(1, 4) -> 1,2,3, doesn't include 4, so we need to do like (1, 4+1), start from 2, because we want to ignore the first row(name)
for row in range(2, sheet.max_row + 1):
    # 7. print(row)
    cell = sheet.cell(row, 3)
    # 8. use cell.value to get the detail value
    print(cell.value)
    # corrected price in discount of 10%
    corrected_price = cell.value * 0.9
    # 9. add a new column
    corrected_price_cell = sheet.cell(row, 4)
    # change the value, corrected price is a float, not a cell
    corrected_price_cell.value = corrected_price

# 11. add a chart, in openpyxl package, we have a module called chart, we import two classes.
# select values from 4th column
from openpyxl.chart import BarChart, Reference

# we use Reference class first, and create a values object
values = Reference(sheet,
          min_row=2,
          max_row=sheet.max_row,
          min_col=4,
          max_col=4)

# use Chart
# 1. object for our chart
chart = BarChart()
# 2. give him the value
chart.add_data(values)
# 3. print the chart
sheet.add_chart(chart, 'e2')

# 10. we need to save it, save it in a new file which can avoid the overwrite the original file
wb.save('transactions2.xlsx')





